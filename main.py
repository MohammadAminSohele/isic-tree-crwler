from bs4 import BeautifulSoup
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


# -----------------------------
# خواندن فایل HTML
# -----------------------------
with open("clicked_html.html", "r", encoding="utf-8") as f:
    html_content = f.read()

soup = BeautifulSoup(html_content, "html.parser")

# پیدا کردن همه پنل‌ها
panels = soup.find_all(
    "div",
    {"id": lambda x: x and x.startswith("jsPanel-")}
)

print(f"🔍 تعداد پنل‌های پیدا شده: {len(panels)}")


# -----------------------------
# استخراج اطلاعات هر محصول
# -----------------------------
def extract_product_info(panel):

    details = panel.find("div", class_="viewport_isic_details")
    if not details:
        return None

    def extract_value(label_text):
        label = details.find("div", string=label_text)
        if label:
            value_div = label.find_next("div", class_="c-dir")
            if value_div:
                return value_div.get_text(strip=True)
        return "-"

    title_path = "-"
    title_label = details.find("div", string="عنوان مسیر")

    if title_label:
        title_div = title_label.find_next("div", class_="c-dir")
        if title_div:
            title_path = title_div.get_text(strip=True)

    return {
        "عنوان مسیر": title_path,
        "عنوان فارسی": extract_value("عنوان فارسی:"),
        "عنوان انگلیسی": extract_value("عنوان انگلیسی:"),
        "کد محصول (ISIC)": extract_value("کدمحصول (ISIC):"),
        "کد تعرفه گمرکی": extract_value("کد تعرفه گمرکی:"),
        "واحد سنجش 1": extract_value("واحد سنجش 1:"),
        "واحد سنجش 2": extract_value("واحد سنجش 2:"),
        "رده‌بندی زیستی": extract_value("رده‌بندی زیستی:"),
        "حوزه کاربری": extract_value("حوزه کاربری:"),
        "عنوان بالادستی": extract_value("عنوان بالادستی:"),
        "سطح": extract_value("سطح:"),
        "معاونت": extract_value("معاونت:"),
        "دفتر تخصصی": extract_value("دفتر تخصصی:"),
        "ISIC 3": extract_value("ISIC 3:"),
        "ISIC 4": extract_value("ISIC 4:"),
        "کد 8 رقمی": extract_value("کد 8 رقمی:"),
        "کدرهگیری": extract_value("کدرهگیری:"),
        "تعداد جواز": extract_value("تعداد جواز:"),
        "تعداد پروانه": extract_value("تعداد پروانه:"),
        "تقسیم بندی BEC": extract_value("تقسیم بندی BEC:"),
        "شناسه CPC": extract_value("شناسه CPC:"),
        "SYS": extract_value("SYS:"),
        "Selection": extract_value("Selection:"),
        "تاریخ استخراج": datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # اضافه کردن تاریخ استخراج
    }


# -----------------------------
# استخراج همه محصولات
# -----------------------------
all_products = []

for panel in panels:
    info = extract_product_info(panel)
    if info:
        all_products.append(info)

print(f"✅ تعداد محصولات استخراج شده: {len(all_products)}")


# -----------------------------
# ذخیره JSON
# -----------------------------
with open("products_info.json", "w", encoding="utf-8") as f:
    json.dump(all_products, f, ensure_ascii=False, indent=2)

print("💾 فایل products_info.json ذخیره شد.")


# -----------------------------
# ذخیره Excel با قابلیت افزودن به انتها
# -----------------------------
def save_products_to_excel(products, filename="products_info.xlsx"):
    
    headers = [
        "عنوان مسیر",
        "عنوان فارسی",
        "عنوان انگلیسی",
        "کد محصول (ISIC)",
        "کد تعرفه گمرکی",
        "واحد سنجش 1",
        "واحد سنجش 2",
        "رده‌بندی زیستی",
        "حوزه کاربری",
        "عنوان بالادستی",
        "سطح",
        "معاونت",
        "دفتر تخصصی",
        "ISIC 3",
        "ISIC 4",
        "کد 8 رقمی",
        "کدرهگیری",
        "تعداد جواز",
        "تعداد پروانه",
        "تقسیم بندی BEC",
        "شناسه CPC",
        "SYS",
        "Selection",
        "تاریخ استخراج"  # ستون جدید برای تاریخ
    ]

    # بررسی وجود فایل
    if os.path.exists(filename):
        print(f"📂 فایل {filename} پیدا شد. اضافه کردن به انتهای فایل...")
        wb = load_workbook(filename)
        ws = wb.active
        
        # پیدا کردن آخرین ردیف دارای داده
        last_row = ws.max_row
        
        # اگر فایل خالی است یا فقط هدر دارد، از ردیف 1 شروع کن
        if last_row <= 1:
            start_row = 2
        else:
            start_row = last_row + 1
            
    else:
        print(f"📄 ایجاد فایل جدید {filename}...")
        wb = Workbook()
        ws = wb.active
        ws.title = "اطلاعات محصولات"
        
        # استایل هدر
        header_font = Font(
            name="B Nazanin",
            size=12,
            bold=True,
            color="FFFFFF"
        )
        
        header_fill = PatternFill(
            fill_type="solid",
            start_color="4150B7",
            end_color="4150B7"
        )
        
        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        start_row = 2

    # تعریف استایل‌ها
    data_font = Font(name="B Nazanin", size=11)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    numeric_columns = {
        "کد محصول (ISIC)",
        "کد تعرفه گمرکی",
        "ISIC 3",
        "ISIC 4",
        "کد 8 رقمی",
        "کدرهگیری",
        "تعداد جواز",
        "تعداد پروانه",
        "SYS",
        "Selection"
    }
    
    row_fill = PatternFill(
        fill_type="solid",
        start_color="F8FBFF",
        end_color="F8FBFF"
    )
    
    # اضافه کردن داده‌های جدید
    for row_offset, product in enumerate(products):
        row_num = start_row + row_offset
        
        for col, header in enumerate(headers, 1):
            value = product.get(header, "-")
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = data_font
            
            if header in numeric_columns:
                cell.alignment = center_align
            else:
                cell.alignment = right_align
            
            cell.fill = row_fill
    
    # تنظیم خودکار عرض ستون‌ها (فقط برای ستون‌هایی که جدیدن ایجاد شدن)
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = min(max_length + 5, 60)
    
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    
    wb.save(filename)
    print(f"✅ {len(products)} رکورد جدید به فایل {filename} اضافه شد.")
    print(f"📊 تعداد کل رکوردها در فایل: {ws.max_row - 1}")


save_products_to_excel(all_products)

print("🎉 عملیات با موفقیت انجام شد.")
